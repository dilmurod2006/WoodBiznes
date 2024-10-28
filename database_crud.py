import sqlite3
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter



# Function to auto-adjust column widths
def adjust_column_widths_for_workers(excel_filename):
    workbook = load_workbook(excel_filename)
    sheet = workbook.active
    # Column widths dictionary
    column_widths = {'A': 22,'B': 10,'C': 10,'D': 10,'E': 10,'F': 24,'G': 17}
    # Adjust column widths based on dictionary
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    # Save the workbook
    workbook.save(excel_filename)

def auto_adjust_column_widths(excel_filename):
    workbook = load_workbook(excel_filename)
    sheet = workbook.active

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass

        adjusted_width = max_length + 2  # Add some padding
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_filename)

# connection to database function
def get_connection():
    conn = sqlite3.connect('database.db')
    cur = conn.cursor()
    return conn, cur

# add admin in database function
def add_admin(full_name, tg_id):
    conn, cur = get_connection()
    created = datetime.now().strftime('%d.%m.%Y %H:%M')
    cur.execute('''
        INSERT INTO Admin (full_name, tg_id, created)
        VALUES (?, ?, ?)
    ''', (full_name, tg_id, created))
    conn.commit()
    cur.fetchall()
    conn.close()

# delete admin in database function
def delete_admin(full_name):
    conn, cur = get_connection()
    cur.execute('DELETE FROM Admin WHERE full_name = ?', (full_name,))
    conn.commit()
    cur.fetchall()
    conn.close()

# add worker in database function
def add_worker(full_name):
    conn, cur = get_connection()
    created = datetime.now().strftime('%d.%m.%Y %H:%M')
    cur.execute('''
        INSERT INTO Workers (full_name, is_deleted, created)
        VALUES (?, ?, ?)
    ''', (full_name, False,created))
    conn.commit()
    conn.close()
# add_worker('test1')
# delete worker in database function
def delete_worker(full_name):
    conn, cur = get_connection()
    cur.execute('''
        UPDATE Workers
        SET is_deleted = TRUE
        WHERE full_name = ?
    ''', (full_name,))
    
    conn.commit()
    conn.close()
# delete_worker('test1')

# Function to get or create a worker ID
def get_worker_id(worker_name):
    try:
        conn, cur = get_connection()
        # Check if the worker exists and get status
        cur.execute('SELECT id, is_deleted FROM Workers WHERE full_name = ?', (worker_name,))
        result = cur.fetchone()
        
        if result:
            worker_id, is_deleted = result
            if not is_deleted:
                # print(worker_id)
                return worker_id
            else:
                conn.close()
                return "bu foydalanuvchi o'chirilgan"
        else:
            conn.close()
            return "bu foydalanuvchi topilmadi"
    except Exception as e:
        conn.close()
get_worker_id('Amonov Dilmurod')


def restore_worker(worker_name):
    try:
        conn, cur = get_connection()
        
        # Check if the worker exists and is deleted
        cur.execute('SELECT id FROM Workers WHERE full_name = ? AND is_deleted = 1', (worker_name,))
        result = cur.fetchone()
        
        if result:
            worker_id = result[0]
            # Restore the worker by setting is_deleted to False
            cur.execute('UPDATE Workers SET is_deleted = 0 WHERE id = ?', (worker_id,))
            conn.commit()
            conn.close()
            return f"{worker_name} ishchi tiklandi xozir active holatda"
        else:
            conn.close()
            return f"Bu {worker_name} ishchi allaqachon faol holatda"
    except Exception as e:
        conn.close()

# Function to parse input data for the daily work
def add_daily_work(data):
    lines = data.strip().split('\n')
    current_worker = None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Check if the line is a worker name (assuming names contain alphabetic characters)
        if any(char.isalpha() for char in line) and not any(char.isdigit() for char in line):
            current_worker = line
        else:
            if current_worker:
                dimensions = line.split('*')
                if len(dimensions) == 2:  # Check if the input format is without length and quantity
                    thickness, width = map(float, dimensions)
                    length = 1  # Default length to 1 if not provided
                    quantity = 1  # Default quantity to 1 if not provided
                elif len(dimensions) == 3:  # Check if the input format is without quantity
                    thickness, width, length = map(float, dimensions)
                    quantity = 1  # Default quantity to 1 if not provided
                elif len(dimensions) == 4:  # Check if the input format includes quantity
                    thickness, width, length, quantity = map(float, dimensions)
                else:
                    continue  # Invalid line format, skip to the next line

                volume = thickness * width * length * quantity
                worker_id = get_worker_id(current_worker)
                if isinstance(worker_id, int):
                    # Connect to the database
                    conn, cur = get_connection()
                    # Insert the work data
                    cur.execute('''
                        INSERT INTO Daily_Work_Data (worker_id, thickness, width, length, quantity, volume_wood, date)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (worker_id, thickness, width, length, quantity, volume, datetime.now().strftime("%d.%m.%Y")))

                    conn.commit()
                    conn.close()
                else:
                    print(worker_id)  # Print error message if worker is not found or deleted

def parse_input_data1(data):
    lines = data.strip().split('\n')
    current_worker = None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.isalpha():  # Check if the line is a worker name
            current_worker = line
        else:
            if current_worker:
                thickness, width, length = map(float, line.split('*'))
                volume = thickness * width * length
                worker_id = get_worker_id(current_worker)
                
                conn = sqlite3.connect('DATABASE1.db')
                cur = conn.cursor()
                
                # Insert the work data
                cur.execute('''
                    INSERT INTO Daily_Work_Data (worker_id, thickness, width, length, volume_wood, description, date)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (worker_id, thickness, width, length, volume, "No description", datetime.now().strftime("%d.%m.%Y %H:%M")))
                
                conn.commit()
                conn.close()

# genarate token
def generate_token(name_wood, truck_number):
    conn, cur = get_connection()
    # Generate the token string
    cur.execute("SELECT MAX(id) FROM WoodsData")
    max_id = cur.fetchone()[0]
    if max_id is None:
        max_id = 1
    else:
        max_id += 1
    token = f"{max_id}-{name_wood}-{truck_number}"
    return token

# add wood data
def add_wood_data(input_data, author_id):
    conn, cur = get_connection()
    lines = input_data.strip().split('\n')
    truck_number = lines[0].strip()
    name_wood = lines[2].strip()
    description = ""
    volume_wood = 0

    # Find the description line
    for i, line in enumerate(lines):
        if line.startswith('#'):
            description = line.strip().lstrip('#').strip()
            break

    token = generate_token(name_wood, truck_number)
    created = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Insert into WoodsData
    cur.execute('''
        INSERT INTO WoodsData (token, author, name_wood, truck_number, description, volume_wood, created)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (token, author_id, name_wood, truck_number, description, volume_wood, created))
    conn.commit()

    # Get the token_id
    cur.execute("SELECT id FROM WoodsData WHERE token = ?", (token,))
    token_id = cur.fetchone()[0]

    # Insert dimensions into Woodsresize
    for line in lines[3:]:
        if line.startswith('#') or not line.strip():
            continue
        thickness, width, length, quantity = map(float, line.split('*'))
        volume = thickness * width * length * quantity
        cur.execute('''
            INSERT INTO Woodsresize (token_id, thickness, width, length, quantity, volume)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (token_id, thickness, width, length, quantity, volume))
        volume_wood += volume

    # Update volume_wood in WoodsData
    cur.execute('''
        UPDATE WoodsData
        SET volume_wood = ?
        WHERE id = ?
    ''', (volume_wood, token_id))
    conn.commit()

# daily workers 1 month report function
def export_daily_work_to_excel(start_date, end_date):
    conn = sqlite3.connect('DATABASE.db')
    cur = conn.cursor()

    # Query to get the daily work data based on the date range
    query = '''
        SELECT
            Workers.full_name,
            Daily_Work_Data.thickness,
            Daily_Work_Data.width,
            Daily_Work_Data.length,
            Daily_Work_Data.quantity,
            Daily_Work_Data.volume_wood,
            Daily_Work_Data.date
        FROM
            Daily_Work_Data
        JOIN
            Workers ON Daily_Work_Data.worker_id = Workers.id
        WHERE Daily_Work_Data.date BETWEEN ? AND ?
    '''
    cur.execute(query, (start_date, end_date))
    rows = cur.fetchall()

    # Create a DataFrame from the query results
    df = pd.DataFrame(rows, columns=['Worker Name', 'Thickness', 'Width', 'Length', 'Quantity', 'Объем древесины', 'Дата'])

    # Format the 'Объем древесины' column to include "m³"
    df['Объем древесины'] = df['Объем древесины'].apply(lambda x: f'{x} m³')

    # Export the DataFrame to an Excel file
    excel_filename = f'работники_отчет_{start_date}_до_{end_date}.xlsx'
    df.to_excel(excel_filename, index=False)

    # Auto-adjust the column widths
    adjust_column_widths_for_workers(excel_filename)
    # send excel file using bot

    # print(f'Data successfully exported to {excel_filename}')

    conn.close()
# export_daily_work_to_excel("01.07.2024", "16.07.2024")

def get_data_from_db(start_date, end_date):
    # SQLite database bilan bog'lanamiz
    conn = sqlite3.connect('DATABASE.db')
    cur = conn.cursor()

    # Ma'lumotlarni olish
    cur.execute('''
        SELECT id, name_wood, truck_number, description, volume_wood, created
        FROM WoodsData
        WHERE created BETWEEN ? AND ?
    ''', (start_date, end_date))
    woods_data = cur.fetchall()

    data = {}
    resize = []

    for row in woods_data:
        wood_id = row[0]
        data[wood_id] = {
            'name_wood': row[1],
            'truck_number': row[2],
            'description': row[3],
            'volume_wood': row[4],
            'created': row[5]
        }

        # Woodsresize dan tegishli resizelarni olish
        cur.execute('''
            SELECT thickness, width, length, quantity, volume
            FROM Woodsresize
            WHERE token_id = ?
        ''', (wood_id,))
        resizes = cur.fetchall()

        for resize_row in resizes:
            resize.append({
                'token_id': wood_id,
                'thickness': resize_row[0],
                'width': resize_row[1],
                'length': resize_row[2],
                'quantity': resize_row[3],
                'volume': resize_row[4]
            })

    # Bog'lanishni yopamiz
    conn.close()

    return data, resize

def get_reports_wood(start_date, end_date):
    # Ma'lumotlarni olish
    data, resize = get_data_from_db(start_date, end_date)

    # Openpyxl bilan ishlaymiz
    wb = Workbook()
    ws = wb.active

    for wood_id, wood_data in data.items():
        # Sarlavhalarni qo'shamiz
        ws.append(['ЭOOOКонтинетУрал П'])
        ws.append(['Номер авто', wood_data['truck_number']])
        ws.append(['Содержимое', wood_data['name_wood']])
        ws.append(['Толщина', 'Ширина', 'Длина', 'Количество', 'Объем'])

        # Resizelarni qo'shamiz
        for resize_data in resize:
            if resize_data['token_id'] == wood_id:
                ws.append([
                    resize_data['thickness'],
                    resize_data['width'],
                    resize_data['length'],
                    resize_data['quantity'],
                    resize_data['volume']
                ])

        # Jami qiymatni hisoblash
        total_volume = sum([r['volume'] for r in resize if r['token_id'] == wood_id])
        ws.append(['ИТОГО:', '', '', '', total_volume])

        #  Add description
        ws.append([wood_data['description']])

        # Bo'sh qator qo'shamiz
        ws.append([])

    # Excel faylini saqlaymiz
    excel_path = f'деревья_отчет_{start_date}_до_{end_date}.xlsx'
    wb.save(excel_path)

    auto_adjust_column_widths(excel_path)

    # print("Ma'lumotlar muvaffaqiyatli Excel fayliga yozildi!")

# Funksiyani chaqiramiz
# start_date = '01.07.2024'
# end_date = '31.07.2024'
# get_reports_wood(start_date, end_date)

def format_date(date):
    return date.strftime("%d.%m.%Y")

def get_admin_info_by_tg_id(tg_id):
    conn, cur = get_connection()
    
    # Admin jadvalidan tg_id bo'yicha id va full_name ni olish
    cur.execute('SELECT id, full_name FROM Admin WHERE tg_id = ?', (tg_id,))
    result = cur.fetchone()
    
    # Ma'lumotlar bazasini yopish
    conn.close()
    
    if result:
        return {"id": result[0], "full_name": result[1]}
    else:
        return None